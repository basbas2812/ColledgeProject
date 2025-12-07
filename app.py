from datetime import datetime, timedelta
import json
from multiprocessing import connection
import uuid
from werkzeug.utils import secure_filename
import requests
from database import engine, Base
from model import User, Expert, Plant, Planting, Disease, Medicine, Advice, Consultation, Treatment
from flask import Flask, jsonify, render_template, request, redirect, session, url_for, flash
import mysql.connector
from mysql.connector import Error
import numpy as np
import os
import cv2
import pandas as pd
import joblib
from skimage.feature import hog
from matplotlib import pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn import svm
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn import metrics
from sklearn.metrics import classification_report 
import random

db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="1234",
    database="myproject"
)

Base.metadata.create_all(bind=engine)

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = 'your_secret_key'  # ควรเปลี่ยนเป็น random string ที่ปลอดภัยกว่า

# ตั้งค่า session lifetime
app.permanent_session_lifetime = timedelta(hours=2)

# ตั้งค่า upload folder
app.config['UPLOAD_FOLDER'] = 'static/uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 🔧 เพิ่ม context processor เพื่อให้ template เข้าถึง session ได้
@app.context_processor
def inject_session():
    return dict(session=session)

@app.route('/static/<path:filename>')
def static_files(filename):
    return app.send_static_file(filename)

@app.route('/')
def index():
    return redirect(url_for('home'))

@app.route('/home')
def home():
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="1234",
        database="myproject"
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT plantId, plantName, description, picture FROM plant ORDER BY plantId DESC")
    plants = cursor.fetchall()
    cursor.close()
    conn.close()

    for p in plants:
        if p["picture"]:
            images = p["picture"].split("|")
            filename = random.choice(images).strip()
        else:
            filename = "uploads/no_image.png"

        # ✅ ให้เหมือน home_search
        p["image_url"] = url_for('static', filename=filename)

    return render_template('home.html', plants=plants)



from flask import url_for

@app.route("/home_search")
def home_search():
    search = request.args.get("search", "")
    plantType = request.args.get("plantType", "")

    conn = mysql.connector.connect(
        host="localhost", user="root", password="1234", database="myproject"
    )
    cursor = conn.cursor(dictionary=True)

    query = "SELECT plantId, plantName, description, picture FROM plant WHERE 1=1"
    params = []

    if search:
        query += " AND plantName LIKE %s"
        params.append(f"%{search}%")

    if plantType:
        query += " AND plantType = %s"
        params.append(plantType)

    cursor.execute(query, params)
    plants = cursor.fetchall()
    cursor.close()
    conn.close()

    for p in plants:
        if p["picture"]:
            images = p["picture"].split("|")
            filename = random.choice(images).strip()
        else:
            filename = "uploads/no_image.png"

        # ✅ ให้ url_for คืนค่า relative path (ไม่ต้อง absolute)
        p["image_url"] = url_for('static', filename=filename, _external=False)


    return jsonify(plants)



@app.route('/register')
def register():
    return render_template('register.html')

@app.route('/login')
def login():
    # ถ้าล็อกอินแล้วให้ไปหน้า home
    if session.get('logged_in'):
        return redirect(url_for('home'))
    return render_template('login.html')


from collections import defaultdict
import json
import mysql.connector

def get_user_consultation_data(user_id):
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="1234",
        database="myproject"
    )
    cursor = conn.cursor(dictionary=True, buffered=True)

    query = """
        SELECT 
            u.userId,
            u.username,
            c.consultationId,
            c.MResult,
            c.dateTime AS consultationDate,
            c.image AS consultationImage,
            c.message AS consultationMessage,
            c.status AS consultationStatus,
            
            a.adviceId,
            a.message AS adviceMessage,
            a.adDateTime,
            
            e.expertId,
            e.expertName
        FROM user u
        JOIN consultation c ON u.userId = c.userId
        LEFT JOIN advice a ON c.consultationId = a.consultationId
        LEFT JOIN expert e ON a.expertId = e.expertId
        WHERE u.userId = %s
        ORDER BY c.dateTime DESC, a.adDateTime DESC
    """

    cursor.execute(query, (user_id,))
    suglist = cursor.fetchall()

    cursor.close()
    conn.close()

    for row in suglist:
        try:
            raw_result = row.get("MResult")
            if raw_result:
                mresult = json.loads(raw_result)
                row["sortedResults"] = sorted(
                    mresult.get("ผลลัพธ์แต่ละภาพ", []),
                    key=lambda x: x.get("เปอร์เซ็นต์ความถูกต้อง", 0),
                    reverse=True
                )
                row["avgConfidence"] = mresult.get("ค่าเฉลี่ยความถูกต้อง", "ไม่ระบุ")
                raw_summary = mresult.get("สรุปผลรวม", "ไม่สามารถสรุปได้")
                if isinstance(raw_summary, list):
                    row["summary"] = " , ".join(raw_summary)
                else:
                    row["summary"] = raw_summary

                # ✅ จัดกลุ่มตามโรค
                grouped = defaultdict(list)
                for item in row["sortedResults"]:
                    grouped[item["โรค"]].append(item)

                groups = [
                    {"disease": disease, "items": items, "count": len(items)}
                    for disease, items in grouped.items()
                ]
                groups.sort(key=lambda g: g["count"], reverse=True)

                row["sortedGroups"] = groups
            else:
                row["sortedResults"] = []
                row["avgConfidence"] = "ไม่ระบุ"
                row["summary"] = "ไม่สามารถสรุปได้"
                row["sortedGroups"] = []
        except Exception as e:
            print("❌ Error parsing MResult:", e)
            row["sortedResults"] = []
            row["avgConfidence"] = "ไม่ระบุ"
            row["summary"] = "ไม่สามารถสรุปได้"
            row["sortedGroups"] = []

    return suglist



@app.route('/view_suggest/<user_id>')
def view_suggest_by_id(user_id):
    current_user_id = session.get('user_id')
    user_type = session.get('user_type')
    
    if not current_user_id:
        flash('กรุณาเข้าสู่ระบบก่อน', 'error')
        return redirect(url_for('login'))
    
    # ✅ ให้ user ดูได้เฉพาะข้อมูลตัวเอง, expert ดูได้ทุกคน
    if user_type == 'user' and user_id != current_user_id:
        flash('คุณสามารถดูได้เฉพาะข้อมูลของตัวเองเท่านั้น', 'warning')
        user_id = current_user_id  # ใช้ string ตรง ๆ
    
    data = get_user_consultation_data(user_id)
    return render_template('viewsuggest.html', consultations=data)




# ---------------- unanswered consultation ----------------
from collections import defaultdict
import json
import mysql.connector

def get_unanswered_consultations(expert_id):
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="1234",
        database="myproject"
    )
    cursor = conn.cursor(dictionary=True, buffered=True)

    query = """
        SELECT 
            c.consultationId,
            c.Mresult,
            c.dateTime AS consultationDate,
            c.image AS consultationImage,
            c.message AS consultationMessage,
            u.userId,
            u.username,
            a.adviceId,
            a.message AS adviceMessage,
            a.expertId
        FROM consultation c
        JOIN user u ON c.userId = u.userId
        JOIN advice a ON c.consultationId = a.consultationId 
            AND a.expertId = %s
        WHERE (a.message IS NULL OR a.message = '')
        ORDER BY c.dateTime DESC;
    """
    cursor.execute(query, (expert_id,))
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    # ✅ แปลง Mresult JSON และเพิ่ม key ที่ใช้ใน template
    for row in results:
        try:
            mresult = json.loads(row["Mresult"])
            row["sortedResults"] = sorted(
                mresult.get("ผลลัพธ์แต่ละภาพ", []),
                key=lambda x: x.get("เปอร์เซ็นต์ความถูกต้อง", 0),
                reverse=True
            )
            row["avgConfidence"] = mresult.get("ค่าเฉลี่ยความถูกต้อง", "ไม่ระบุ")
            raw_summary = mresult.get("สรุปผลรวม", "ไม่สามารถสรุปได้")

            # ✅ ถ้าเป็น list → รวมเป็น string
            if isinstance(raw_summary, list):
                row["summary"] = " , ".join(raw_summary)
            else:
                row["summary"] = raw_summary

            # ✅ จัดกลุ่มตามโรคและเรียงตามจำนวนรูป
            grouped = defaultdict(list)
            for item in row["sortedResults"]:
                grouped[item["โรค"]].append(item)

            groups = [
                {"disease": disease, "items": items, "count": len(items)}
                for disease, items in grouped.items()
            ]
            groups.sort(key=lambda g: g["count"], reverse=True)

            row["sortedGroups"] = groups

        except Exception:
            row["sortedResults"] = []
            row["avgConfidence"] = "ไม่ระบุ"
            row["summary"] = "ไม่สามารถสรุปได้"
            row["sortedGroups"] = []

    return results


@app.route('/list_suggest')
def list_suggest():
    if session.get('user_type') != 'expert':
        flash('เฉพาะผู้เชี่ยวชาญเท่านั้นที่สามารถเข้าใช้ฟีเจอร์นี้ได้', 'error')
        return redirect(url_for('home'))

    expert_id = session.get('expert_id')
    data = get_unanswered_consultations(expert_id)

    # ✅ เรียงตามวันที่ล่าสุด
    data.sort(key=lambda x: x["consultationDate"], reverse=True)

    return render_template('listsuggest.html', consultations=data)





# ---------------- give suggest ----------------
@app.route('/add_advice_simple/<adviceid>', methods=['POST'])
def add_advice_simple(adviceid):
    if session.get('user_type') != 'expert':
        flash('เฉพาะผู้เชี่ยวชาญเท่านั้นที่สามารถตอบได้', 'error')
        return redirect(url_for('home'))

    expert_id = session.get('expert_id')
    message = request.form.get('message')

    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="1234",
        database="myproject"
    )
    cursor = conn.cursor()

    # ✅ ตรวจสอบว่า adviceId นี้เป็นของ expert ที่ login อยู่จริง
    cursor.execute("SELECT consultationId FROM advice WHERE adviceId = %s AND expertId = %s", (adviceid, expert_id))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        flash('❌ คุณไม่มีสิทธิ์ตอบคำปรึกษานี้', 'error')
        return redirect(url_for('list_suggest'))

    consultationId = row[0]

    # 🔹 อัปเดตคำตอบ
    cursor.execute("""
        UPDATE advice
        SET message = %s, adDateTime = NOW()
        WHERE adviceId = %s AND expertId = %s
    """, (message, adviceid, expert_id))

    # 🔹 อัปเดตสถานะ consultation
    cursor.execute("""
        UPDATE consultation
        SET status = %s
        WHERE consultationId = %s
    """, ("ตอบแล้ว", consultationId))

    conn.commit()
    cursor.close()
    conn.close()

    flash('✅ ตอบคำปรึกษาสำเร็จ และอัปเดตสถานะเรียบร้อย', 'success')
    return redirect(url_for('list_suggest'))





# ---------------- view plant ----------------
@app.route('/view_plant/<string:plant_id>')
def view_plant(plant_id):
    conn = None
    cursor = None
    
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="1234",
            database="myproject"
        )
        cursor = conn.cursor(dictionary=True, buffered=True)

        # ✅ ดึงข้อมูลพืชและวิธีปลูก
        cursor.execute("""   
            SELECT plant.picture, plant.plantName, plant.prepare, plant.care, planting.plantingMethod
            FROM plant
            INNER JOIN planting ON plant.plantId = planting.plantId
            WHERE plant.plantId = %s
        """, (plant_id,))
        rows = cursor.fetchall()

        if not rows:
            flash('ไม่พบข้อมูลพืช', 'error')
            return redirect(url_for('home'))

        plant = {
            'picture': rows[0]['picture'],
            'plantName': rows[0]['plantName'],
            'prepare': rows[0]['prepare'],
            'care': rows[0]['care']
        }
        plant_methods = [{'plantingMethod': row['plantingMethod']} for row in rows]

        # ✅ ดึงข้อมูลโรค + วิธีรักษา + ยา
        cursor.execute("""
            SELECT 
                d.diseaseId, d.diseaseName, d.symptoms,
                t.treatmentMethods,
                m.medicineName, m.dosage, m.quantityType
            FROM disease d
            LEFT JOIN treatment t ON d.diseaseId = t.diseaseId
            LEFT JOIN medicine m ON d.diseaseId = m.diseaseId
            WHERE d.plantId = %s
            ORDER BY d.diseaseId
        """, (plant_id,))
        disease_rows = cursor.fetchall()

        # ✅ รวมข้อมูลโรคแบบกลุ่ม
        diseases = {}
        for row in disease_rows:
            disease_id = row['diseaseId']
            if disease_id not in diseases:
                diseases[disease_id] = {
                    'diseaseName': row['diseaseName'],
                    'symptoms': row['symptoms'],
                    'treatments': [],   # เก็บเป็น list
                    'medicines': []
                }
            # ✅ เก็บ treatmentMethods หลายค่า
            if row['treatmentMethods'] and row['treatmentMethods'] not in diseases[disease_id]['treatments']:
                diseases[disease_id]['treatments'].append(row['treatmentMethods'])

            # ✅ เก็บ medicines
            if row['medicineName']:
                dose_text = f"{row['dosage']} {row['quantityType']}" if row['quantityType'] else row['dosage']
                med_item = {'medicineName': row['medicineName'], 'dosage': dose_text}
                if med_item not in diseases[disease_id]['medicines']:
                    diseases[disease_id]['medicines'].append(med_item)


        return render_template(
            'viewplant.html',
            plant=plant,
            methods=plant_methods,
            diseases=list(diseases.values())
        )

    except Exception as e:
        print(f"Database error: {e}")
        flash('เกิดข้อผิดพลาดในการโหลดข้อมูล', 'error')
        return redirect(url_for('home'))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



# ---------------- register user ----------------
@app.route('/register_submit', methods=['POST'])
def register_submit():
    try:
        name = request.form['username']
        password = request.form['password']
        email = request.form['email']
        address = request.form['address']

        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT email FROM user WHERE email = %s", (email,))
        if cursor.fetchone():
            flash('อีเมลนี้มีการใช้งานแล้ว', 'error')
            cursor.close()
            return render_template('register.html')

        cursor.execute("SELECT email FROM expert WHERE email = %s", (email,))
        if cursor.fetchone():
            flash('อีเมลนี้มีการใช้งานแล้ว', 'error')
            cursor.close()
            return render_template('register.html')

        cursor.execute("SELECT MAX(CAST(SUBSTRING(userId, 2) AS UNSIGNED)) FROM user WHERE userId LIKE 'U%'")
        result = cursor.fetchone()
        next_id = (result[0] or 0) + 1
        auto_genId = f"U{next_id:04d}"

        # ✅ บันทึกข้อมูล
        sql = "INSERT INTO user (userId, userName, password, email, address) VALUES (%s, %s, %s, %s, %s)"
        values = (auto_genId, name, password, email, address)

        cursor.execute(sql, values)
        db.commit()
        cursor.close()

        flash('สมัครสมาชิกสำเร็จ กรุณาเข้าสู่ระบบ', 'success')
        return redirect(url_for('login'))

    except Error as e:
        flash(f'เกิดข้อผิดพลาด: {e}', 'error')
        return render_template('register.html')



# ---------------- login ----------------
@app.route('/login_submit', methods=['POST'])
def login_submit():
    conn = None
    cursor = None
    try:
        email = request.form.get('email')
        password = request.form.get('password')
        
        if not email or not password:
            flash('กรุณากรอกอีเมลและรหัสผ่าน', 'error')
            return render_template('login.html')
        
        conn = mysql.connector.connect(
            host="localhost", user="root", password="1234", database="myproject"
        )
        cursor = conn.cursor(dictionary=True)

        # เช็ค user
        cursor.execute("SELECT * FROM user WHERE email = %s", (email,))
        user = cursor.fetchone()

        if user and user['password'] == password:
            session['user_id'] = user['userId']
            session['user_name'] = user['username']
            session['user_email'] = user['email']
            session['user_address'] = user['address']
            session['user_type'] = "user"
            session['logged_in'] = True
            session.permanent = True
            flash(f'เข้าสู่ระบบสำเร็จ ยินดีต้อนรับ {user["username"]}', 'success')
            return redirect(url_for('home'))

        # ถ้าไม่เจอใน user → เช็ค expert
        cursor.execute("SELECT * FROM expert WHERE email = %s", (email,))
        expert = cursor.fetchone()

        if expert and expert['password'] == password:
            session['user_id'] = expert['expertId']
            session['expert_id'] = expert['expertId']
            session['user_name'] = expert['expertName']
            session['user_email'] = expert['email']
            session['user_address'] = expert['address']
            session['user_type'] = "expert"
            session['logged_in'] = True
            session.permanent = True
            flash(f'เข้าสู่ระบบสำเร็จ ยินดีต้อนรับ {expert["expertName"]}', 'success')
            return redirect(url_for('home'))

        flash('อีเมลหรือรหัสผ่านไม่ถูกต้อง', 'error')
        return render_template('login.html')

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/logout')
def logout():
    session.clear()
    flash('ออกจากระบบเรียบร้อยแล้ว', 'success')
    return redirect(url_for('home'))   # ✅ กลับไปหน้า home เห็นภาพได้แม้ไม่ login


# 🔧 ปรับ before_request ให้เหมาะสม
@app.before_request
def check_session():
    """
    ตรวจสอบ session ก่อนเข้าถึงทุก route
    ยกเว้น public routes และ static files
    """
    # Public routes ที่ไม่ต้อง login
    public_routes = [
        'static',           # Flask built-in static handler
        'login', 
        'login_submit', 
        'register', 
        'register_submit', 
        'home', 
        'home_search',
        'index', 
        'view_plant'
    ]
    
    # ✅ ยกเว้น public routes และ static files
    if request.endpoint in public_routes or request.path.startswith('/static/'):
        return None  # ปล่อยผ่าน
    
    # ✅ เช็ค login สำหรับ routes อื่นๆ
    if not session.get('logged_in'):
        flash('กรุณาเข้าสู่ระบบก่อน', 'error')
        return redirect(url_for('login'))
    

@app.route('/editprofile')
def editprofile():
    if not session.get('logged_in'):
        flash('กรุณาเข้าสู่ระบบก่อน', 'error')
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_type = session.get('user_type', 'user')

    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="1234",
            database="myproject"
        )
        cursor = conn.cursor(dictionary=True)

        table_name = "expert" if user_type == "expert" else "user"
        id_field = "expertId" if user_type == "expert" else "userId"

        cursor.execute(f"SELECT * FROM {table_name} WHERE {id_field} = %s", (user_id,))
        user = cursor.fetchone()
        if user:
            user_data = {
                'user_id': user[id_field],
                'user_name': user['expertName'] if user_type == 'expert' else user['username'],
                'user_email': user['email'],
                'user_password': user['password'],
                'user_address': user['address']
            }
            return render_template('editprofile.html', user=user_data)
        else:
            flash('⚠️ ไม่พบข้อมูลผู้ใช้', 'error')
            return redirect(url_for('home'))

    except Exception as e:
        print(f"Error: {e}")
        flash('❌ เกิดข้อผิดพลาดในการโหลดข้อมูล', 'error')
        return redirect(url_for('home'))
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()



# ---------------- edit profile ----------------
@app.route('/edit_profile', methods=['GET', 'POST'])
def edit_profile():
    if not session.get('logged_in'):
        flash('กรุณาเข้าสู่ระบบก่อนใช้งาน', 'error')
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    user_type = session.get('user_type', 'user')
    conn = None
    cursor = None
    
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="1234",
            database="myproject"
        )
        cursor = conn.cursor(dictionary=True)
        
        if request.method == 'POST':
            username = request.form['username']
            email = request.form['email']
            password = request.form['password']
            address = request.form['address']
            
            # เลือกตารางตาม user_type
            table_name = "expert" if user_type == "expert" else "user"
            id_field = "expertId" if user_type == "expert" else "userId"
            name_field = "expertName" if user_type == "expert" else "username"
            
            sql = f"""
                UPDATE {table_name} 
                SET {name_field}=%s, email=%s, password=%s, address=%s 
                WHERE {id_field}=%s
            """
            values = (username, email, password, address, user_id)
            
            cursor.execute(sql, values)
            conn.commit()

            # อัปเดต session
            session['user_name'] = username
            session['user_email'] = email
            session['user_password'] = password
            session['user_address'] = address
            
            flash('✅ อัปเดตโปรไฟล์เรียบร้อยแล้ว', 'success')
            return redirect(url_for('editprofile'))
        
        else:
            # เลือกตารางตาม user_type
            table_name = "expert" if user_type == "expert" else "user"
            id_field = "expertId" if user_type == "expert" else "userId"
            
            cursor.execute(f"SELECT * FROM {table_name} WHERE {id_field} = %s", (user_id,))
            user = cursor.fetchone()
            if user:
                # ✅ map key ให้ตรงกับ template
                user_data = {
                    'user_id': user[id_field],
                    'user_name': user['expertName'] if user_type == 'expert' else user['username'],
                    'user_email': user['email'],
                    'user_password': user['password'],
                    'user_address': user['address']
                }
                return render_template('editprofile.html', user=user_data)

            else:
                flash('⚠️ ไม่พบข้อมูลผู้ใช้', 'error')
                return redirect(url_for('home'))
                
    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        flash('❌ เกิดข้อผิดพลาดในการเชื่อมต่อฐานข้อมูล', 'error')
        return redirect(url_for('home'))
    except Exception as e:
        print(f"เกิดข้อผิดพลาด: {e}")
        flash('❌ เกิดข้อผิดพลาดในการโหลดข้อมูล', 'error')
        return redirect(url_for('home'))
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()


# ---------------- add plant page ----------------
@app.route('/add_plant_page')
def add_plant_page():
    # เช็คสิทธิ์ผู้เชี่ยวชาญ
    if session.get('user_type') != 'expert':
        flash('เฉพาะผู้เชี่ยวชาญเท่านั้นที่สามารถเข้าใช้ฟีเจอร์นี้ได้', 'error')
        return redirect(url_for('home'))
    return render_template('addplant.html')

# ---------------- upload plant ----------------
from werkzeug.utils import secure_filename
import os
import mysql.connector

UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/add_submit_plant', methods=['POST'])
def add_submit_plant():
    cursor = None
    try:
        # ----------------------------
        # 1. ตรวจสอบสิทธิ์
        # ----------------------------
        if session.get('user_type') != 'expert':
            flash('เฉพาะผู้เชี่ยวชาญเท่านั้นที่สามารถเพิ่มข้อมูลพืชได้', 'error')
            return redirect(url_for('home'))
        
        # ----------------------------
        # 2. Upload รูปภาพ
        # ----------------------------
        try:
            plant_images = request.files.getlist('plantImage')
            saved_paths = []
            for plantimg in plant_images:
                if plantimg and allowed_file(plantimg.filename):
                    filename = secure_filename(plantimg.filename)
                    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    plantimg.save(save_path)
                    saved_paths.append(f"uploads/{filename}")
            plantimg_path = "|".join(saved_paths) if saved_paths else None
        except Exception as e:
            flash(f'อัปโหลดรูปภาพผิดพลาด: {str(e)}', 'error')
            return redirect(url_for('add_plant_page'))

        # ----------------------------
        # 3. รับค่าจากฟอร์ม
        # ----------------------------
        plantname = request.form.get('plantName', '')
        plantcare = request.form.get('careInstructions', '')
        plantprepare = request.form.get('plantPrepare', '')
        plant_type = request.form.get('plantType', '')
        description = request.form.get('description', '')

        planting_methods = request.form.getlist('plantingMethods[]')
        disease_names = request.form.getlist('diseaseNames[]')
        disease_symptoms = request.form.getlist('diseaseSymptoms[]')

        if not plantname or not plantcare:
            flash('กรุณากรอกชื่อพืชและวิธีการดูแล', 'error')
            return redirect(url_for('add_plant_page'))

        cursor = db.cursor()

        # ----------------------------
        # 4. ตรวจสอบชื่อซ้ำ
        # ----------------------------
        cursor.execute("SELECT COUNT(*) FROM plant WHERE plantName = %s", (plantname,))
        exists = cursor.fetchone()[0]
        if exists > 0:
            return jsonify({
                "success": False,
                "message": "มีชื่อพืชนี้อยู่แล้วในระบบ กรุณาใช้ชื่ออื่น"
            }), 400

        # ----------------------------
        # 5. Insert Plant
        # ----------------------------
        try:
            cursor.execute("SELECT MAX(CAST(SUBSTRING(plantId, 2) AS UNSIGNED)) FROM plant WHERE plantId LIKE 'P%'")
            result = cursor.fetchone()
            next_plant_id = (result[0] or 0) + 1
            plant_id = f"P{next_plant_id:04d}"

            cursor.execute("""
                INSERT INTO plant (plantId, care, description, picture, plantName, plantType, prepare, consultation_consultationId)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NULL)
            """, (plant_id, plantcare, description, plantimg_path, plantname, plant_type, plantprepare))
        except Exception as e:
            db.rollback()
            flash(f'❌ Error ตอนเพิ่มข้อมูลพืช: {str(e)}', 'error')
            return redirect(url_for('add_plant_page'))

        # ----------------------------
        # 6. Insert Planting Methods
        # ----------------------------
        try:
            if planting_methods and any(method.strip() for method in planting_methods):
                for method in planting_methods:
                    if method.strip():
                        cursor.execute("SELECT MAX(CAST(SUBSTRING(plantingId, 3) AS UNSIGNED)) FROM planting WHERE plantingId LIKE 'PT%'")
                        result = cursor.fetchone()
                        next_planting_id = (result[0] or 0) + 1
                        planting_id = f"PT{next_planting_id:04d}"

                        cursor.execute("""
                            INSERT INTO planting (plantingId, plantingMethod, plantId)
                            VALUES (%s, %s, %s)
                        """, (planting_id, method.strip(), plant_id))
        except Exception as e:
            flash(f'เพิ่มข้อมูลวิธีปลูกผิดพลาด: {str(e)}', 'error')
            db.rollback()
            return redirect(url_for('add_plant_page'))

        # ----------------------------
        # 7. Insert Diseases, Treatments, Medicines
        # ----------------------------
        try:
            if disease_names and any(name.strip() for name in disease_names):
                for i, disease_name in enumerate(disease_names):
                    if disease_name.strip():
                        # Disease
                        cursor.execute("SELECT MAX(CAST(SUBSTRING(diseaseId, 2) AS UNSIGNED)) FROM disease WHERE diseaseId LIKE 'D%'")
                        result = cursor.fetchone()
                        next_disease_id = (result[0] or 0) + 1
                        disease_id = f"D{next_disease_id:04d}"

                        symptom = disease_symptoms[i] if i < len(disease_symptoms) else ""
                        try:
                            cursor.execute("""
                                INSERT INTO disease (diseaseId, diseaseName, symptoms, plantId)
                                VALUES (%s, %s, %s, %s)
                            """, (disease_id, disease_name.strip(), symptom.strip(), plant_id))
                        except Exception as e:
                            db.rollback()
                            flash(f'❌ Error ตอนเพิ่มข้อมูลโรค {disease_name}: {str(e)}', 'error')
                            return redirect(url_for('add_plant_page'))
                        
                        # Treatments
                        treatment_methods = request.form.getlist(f"treatmentMethods[{i+1}][]")
                        for treatment_method in treatment_methods:
                            if treatment_method.strip():
                                cursor.execute("SELECT MAX(CAST(SUBSTRING(treatmentId, 2) AS UNSIGNED)) FROM treatment WHERE treatmentId LIKE 'T%'")
                                result = cursor.fetchone()
                                next_treatment_id = (result[0] or 0) + 1
                                treatment_id = f"T{next_treatment_id:04d}"

                                cursor.execute("""
                                    INSERT INTO treatment (treatmentId, treatmentMethods, diseaseId)
                                    VALUES (%s, %s, %s)
                                """, (treatment_id, treatment_method.strip(), disease_id))

                        # Medicines
                        medicine_names_for_disease = request.form.getlist(f"medicineNames[{i+1}][]")
                        medicine_dosages_for_disease = request.form.getlist(f"medicineDosages[{i+1}][]")
                        medicine_units_for_disease = request.form.getlist(f"medicineUnits[{i+1}][]")

                        for j, med_name in enumerate(medicine_names_for_disease):
                            if med_name.strip():
                                cursor.execute("SELECT MAX(CAST(SUBSTRING(medicineId, 2) AS UNSIGNED)) FROM medicine WHERE medicineId LIKE 'M%'")
                                result = cursor.fetchone()
                                next_medicine_id = (result[0] or 0) + 1
                                medicine_id = f"M{next_medicine_id:04d}"

                                dosage = 0
                                if j < len(medicine_dosages_for_disease):
                                    try:
                                        dosage = float(medicine_dosages_for_disease[j]) if medicine_dosages_for_disease[j].strip() else 0
                                    except Exception:
                                        dosage = 0

                                quantity_type = None
                                if j < len(medicine_units_for_disease):
                                    quantity_type = medicine_units_for_disease[j].strip()
                                try:
                                    cursor.execute("""
                                        INSERT INTO medicine (medicineId, dosage, quantityType, medicineName, diseaseId)
                                        VALUES (%s, %s, %s, %s, %s)
                                    """, (medicine_id, dosage, quantity_type, med_name.strip(), disease_id))
                                except Exception as e:
                                    db.rollback()
                                    flash(f'❌ Error ตอนเพิ่มยาชื่อ {med_name}: {str(e)}', 'error')
                                    return redirect(url_for('add_plant_page'))
        except Exception as e:
            flash(f'เพิ่มข้อมูลโรค/การรักษา/ยา ผิดพลาด: {str(e)}', 'error')
            db.rollback()
            return redirect(url_for('add_plant_page'))

        # ----------------------------
        # 8. Commit
        # ----------------------------
        db.commit()
        flash('เพิ่มข้อมูลพืชเรียบร้อยแล้ว', 'success')
        return redirect(url_for('home'))

    except mysql.connector.Error as e:
        if cursor:
            db.rollback()
        flash(f'เกิดข้อผิดพลาดในฐานข้อมูล: {str(e)}', 'error')
        return redirect(url_for('add_plant_page'))
    except Exception as e:
        if cursor:
            db.rollback()
        flash(f'เกิดข้อผิดพลาดไม่ทราบสาเหตุ: {str(e)}', 'error')
        return redirect(url_for('add_plant_page'))
    finally:
        if cursor:
            cursor.close()


PREDICTOR_URL = "http://127.0.0.1:5001/predict"



@app.route('/analyze')
def analyze():
    # เช็คสิทธิ์ (optional)
    if session.get('user_type') == 'expert':
        flash('ผู้เชี่ยวชาญไม่สามารถใช้ฟีเจอร์นี้ได้', 'warning')
        return redirect(url_for('home'))
    return render_template('analyze.html')



@app.route('/analyze_plant', methods=['POST'])
def analyze_plant():
    try:
        files = request.files.getlist('file')
        plant_type = request.form.get('plantType')

        if not files or len(files) == 0:
            return jsonify({"error": "กรุณาอัปโหลดไฟล์ภาพ"}), 400
        if not plant_type:
            return jsonify({"error": "กรุณาระบุชนิดพืช"}), 400

        # ✅ เตรียมข้อมูลสำหรับส่งไปยัง predictor
        file_payload = []
        for file in files:
            file_payload.append(
                ("file", (file.filename, file.stream, file.mimetype))
            )

        response = requests.post(
            PREDICTOR_URL,
            files=file_payload,
            data={"plantType": plant_type}
        )

        if response.status_code != 200:
            return jsonify({"error": "ไม่สามารถประมวลผลได้"}), 500

        try:
            data = response.json()
        except Exception:
            return jsonify({"error": "ไม่สามารถแปลงข้อมูลจาก Predictor ได้"}), 500

        # ✅ สร้างผลลัพธ์รวม
        results = data.get("ผลลัพธ์แต่ละภาพ", [])
        avg_confidence = data.get("ค่าเฉลี่ยความถูกต้อง", "ไม่ระบุ")

        # ✅ นับจำนวนโรคทั้งหมด
        disease_counts = {}
        for r in results:
            if "โรค" in r:
                disease_name = r["โรค"]
                disease_counts[disease_name] = disease_counts.get(disease_name, 0) + 1

        summary = []
        if disease_counts:
            total = sum(disease_counts.values())
            # ✅ วนลูปทุกโรค เรียงจากมากไปน้อย
            sorted_diseases = sorted(disease_counts.items(), key=lambda x: x[1], reverse=True)
            for disease, count in sorted_diseases:
                percentage = round((count / total) * 100, 2)
                summary.append(f"น่าจะเป็น {disease} ({percentage}%)")

        return jsonify({
            "จำนวนภาพที่วิเคราะห์": len(results),
            "ค่าเฉลี่ยความถูกต้อง": avg_confidence,
            "ผลลัพธ์แต่ละภาพ": results,
            "สรุปผลรวม": summary if summary else ["ไม่สามารถสรุปได้"]
        })

    except Exception as e:
        return jsonify({"error": f"เกิดข้อผิดพลาด: {str(e)}"}), 500




# ---------------- ส่งให้ผู้เชี่ยวชาญคนเดียว ----------------
@app.route("/get_experts")
def get_experts():
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT expertId, expertName FROM expert")
    experts = cursor.fetchall()
    return jsonify(experts)



# ---------------- give consultation ----------------
@app.route("/request_consult", methods=["POST"])
def request_consult():
    try:
        if 'user_id' not in session:
            return jsonify({"error": "กรุณาเข้าสู่ระบบก่อน"}), 401

        userId = session['user_id']
        Mresult = request.form.get("Mresult")
        status = request.form.get("status", "รอคำตอบ")
        expertId = request.form.get("expertId")
        message = request.form.get("message")

        # 🔹 จัดการไฟล์รูปภาพหลายไฟล์
        files = request.files.getlist("file")
        saved_filenames = []

        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(save_path)
                saved_filenames.append(f"uploads/{filename}")

        # รวมชื่อไฟล์เป็น string เดียว (คั่นด้วย |)
        file_paths = "|".join(saved_filenames) if saved_filenames else None

        # เชื่อมต่อ DB
        conn = mysql.connector.connect(
            host="localhost", user="root", password="1234", database="myproject"
        )
        cursor = conn.cursor()

        # consultationId ใหม่
        cursor.execute("SELECT MAX(CAST(SUBSTRING(consultationId,2) AS UNSIGNED)) FROM consultation")
        result = cursor.fetchone()
        next_id = (result[0] or 0) + 1
        consultationId = f"C{next_id:04d}"

        # insert consultation
        sql = """INSERT INTO consultation 
                 (consultationId, Mresult, dateTime, image, message, status, userId)
                 VALUES (%s, %s, %s, %s, %s, %s, %s)"""
        values = (consultationId, Mresult, datetime.now(), file_paths, message, status, userId)
        cursor.execute(sql, values)

        # adviceId ใหม่
        cursor.execute("SELECT MAX(CAST(SUBSTRING(adviceId,2) AS UNSIGNED)) FROM advice")
        result = cursor.fetchone()
        next_advice_id = (result[0] or 0) + 1
        adviceId = f"A{next_advice_id:04d}"

        # insert advice เปล่า
        sql_advice = """INSERT INTO advice 
                        (adviceId, message, adDateTime, consultationId, expertId)
                        VALUES (%s, %s, %s, %s, %s)"""
        values_advice = (adviceId, None, datetime.now(), consultationId, expertId)
        cursor.execute(sql_advice, values_advice)

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": True, "consultationId": consultationId, "adviceId": adviceId})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})



# ---------------- upload expert page ----------------
@app.route('/upload_expert_page')
def upload_expert_page():
    # เช็คสิทธิ์ผู้เชี่ยวชาญ
    if session.get('user_type') != 'expert':
        flash('เฉพาะผู้เชี่ยวชาญเท่านั้นที่สามารถเข้าใช้ฟีเจอร์นี้ได้', 'error')
        return redirect(url_for('home'))
    return render_template('uploadexpert.html')

# ---------------- upload expert ----------------
from openpyxl import load_workbook

@app.route('/upload_expert', methods=['POST'])
def upload_expert():
    cursor = None
    added_experts = []  # ✅ เก็บเฉพาะคนที่เพิ่มใหม่จริง ๆ

    try:
        if 'file' not in request.files:
            flash("กรุณาเลือกไฟล์ก่อน", "error")
            return redirect(url_for('upload_expert_page'))

        file = request.files['file']
        if file.filename == '':
            flash("ไม่ได้เลือกไฟล์", "error")
            return redirect(url_for('upload_expert_page'))

        workbook = load_workbook(file)
        sheet = workbook.active
        cursor = db.cursor()

        for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            if not row or not row[0]:
                continue

            expertName, password, email, address = row

            # ✅ ตรวจสอบอีเมลซ้ำ
            cursor.execute("SELECT expertId FROM expert WHERE email = %s", (email,))
            if cursor.fetchone():
                continue  # ข้ามถ้าอีเมลซ้ำ

            # ✅ สร้าง expertId ใหม่
            cursor.execute("SELECT MAX(CAST(SUBSTRING(expertId, 2) AS UNSIGNED)) FROM expert WHERE expertId LIKE 'E%'")
            result = cursor.fetchone()
            next_id = (result[0] or 0) + 1
            expert_id = f"E{next_id:04d}"

            # ✅ เพิ่มข้อมูล
            cursor.execute("""
                INSERT INTO expert (expertId, expertName, password, email, address)
                VALUES (%s, %s, %s, %s, %s)
            """, (expert_id, str(expertName), str(password), str(email), str(address)))

            added_experts.append(str(expertName))

        db.commit()
        count = len(added_experts)

        if count > 0:
            flash(f"✅ อัปโหลดผู้เชี่ยวชาญสำเร็จ {count} คน", "success")
        else:
            flash("ไม่มีผู้เชี่ยวชาญใหม่ที่ถูกเพิ่ม (อาจมีอีเมลซ้ำ)", "info")

        return render_template("uploadexpert.html", count=count, names=added_experts)

    except Exception as e:
        if cursor:
            db.rollback()
        flash(f"เกิดข้อผิดพลาด: {str(e)}", "error")
        return redirect(url_for('upload_expert_page'))

    finally:
        if cursor:
            cursor.close()


# ----------------- Run -----------------
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
    